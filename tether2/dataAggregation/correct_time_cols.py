""" Module to correct the time columns in v1 of teh DFDRs
stopwatch, vtp and transmitter all encoded as strings and the stopwatch and vtp have multiple 
0:00.000 entries 
"""
import pandas as pd
import numpy as np
import argparse
import time
from openpyxl import load_workbook
from datetime import datetime

def correct_time(excel_file,worksheet,output_file = None):
    # extract data from excel_file
    print(f'Extracting data from: {excel_file}')
    excel_file = open(excel_file,'rb')
    headers = pd.read_excel(excel_file,sheet_name = worksheet,
                    header=None,nrows=3)
    headers = headers.fillna(method='ffill',axis=1)
    data = pd.read_excel(excel_file,sheet_name = worksheet,
                    header=None,skiprows=3)

    # ----------------------- Create corrected times --------------------------------
    print('Creating corrected time values')
    """df_to_secs converts datatime to number of seconds since midnight"""
    def dt_to_secs(pt):   
        return pt.microsecond/1e6 + pt.second + (pt.minute + pt.hour * 60) * 60

    # create unix timestamp from transmitter time
    time = data[4]
    master_time =[]
    for t in time.to_list():
        pt = datetime.strptime(headers[0][1].split()[-1] + t,'(%d/%m/%Y)%H:%M:%S.%f')
        master_time.append(pt.timestamp())

    # fix stopwatch times
    sw_time = data[1]
    start_index = next((i for i, x in enumerate(sw_time) if x>0),None)
    unix_start = master_time[start_index]-sw_time[start_index]
    sw_time = [np.round(mt - unix_start,2) for mt in master_time] 

    # decode vtp time
    vtp_time = data[3]
    vtp_time  = [dt_to_secs(datetime.strptime(t,'%M:%S.%f')) if isinstance(t, str) else np.NaN for t in vtp_time ]
    for i,x in enumerate(vtp_time[:-1]):
        if vtp_time[i+1] == 0:
            vtp_time[i] = np.NaN
    if vtp_time[-2] == np.NaN:
        vtp_time[-1] = np.NaN

    # ----------------------- Save to an excel file --------------------------------
    wb = load_workbook(excel_file)
    ws = wb[worksheet]
    for i, row in enumerate(ws.iter_rows(min_col=2, min_row = 4, max_row=4+len(master_time)-1, max_col=4)):
        new_values = [master_time[i],sw_time[i],vtp_time[i]]
        for c_i, cell in enumerate(row):
            cell.value = new_values[c_i]
    
    # name columns
    ws['B3'].value = 'Unix Time (s)'
    ws['C3'].value = 'Master Stopwatch (s)'
    ws['D3'].value = 'VTP Cam Time (s)'

    output_file = excel_file if output_file is None else output_file
    print(f'Saving to File: {output_file}')
    wb.save(output_file)
    excel_file.close()
    print('Complete')

if __name__ == '__main__':
    ap = argparse.ArgumentParser(description=__doc__,
                            formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("-x", "--excel_file", required = True, help = "excelfile")
    ap.add_argument("-w", "--worksheet", default = 'DFDR', help = "worksheet name")
    ap.add_argument("-o", "--output_file", help = "Output file name - if none will overwrite excel_file")
    args = vars(ap.parse_args())
    correct_time(**args)