import pandas as pd
import numpy as np
import argparse
import time
from openpyxl import load_workbook
from datetime import datetime
from openpyxl import load_workbook
from scipy.interpolate import griddata,interp1d
import pickle

def add_fwt_angles(excel_filename,angle_data_filename,angle_calib_filename,worksheet='DFDR',time_offset=0,is_right=True,output_file = None):
    print('Loading Excel File')
    excel_file = open(excel_filename,'rb')
    data = pd.read_excel(excel_file,sheet_name = worksheet,
                  header=None,skiprows=3,  engine='openpyxl')
    df = get_angle_data(angle_data_filename,is_right)

    print('Loading calibration')
    calib = pickle.load(open(angle_calib_filename,'rb'))
    calib_func = interp1d(calib['angle_act'],calib['angle_calib'],fill_value='extrapolate')

    print('Interpolate data onto dfdr times')
    df['angle_calib'] = calib_func(df['angle'])
    f_angle = interp1d(df['t']+ time_offset,df['angle_calib'],bounds_error=False)


    # ----------------------- Save to an excel file --------------------------------
    col_index = 45 if is_right else 44
    ind = np.logical_and(data[2] >= min(df['t'] + time_offset),data[2] <= max(df['t'] + time_offset))
    wb = load_workbook(excel_file)
    ws = wb[worksheet]
    vals = data.loc[ind,2]
    for i,val in zip(vals.index,vals):
        ws.cell(row = 4 + i,column = col_index).value = f_angle([val])[0]

    output_file = excel_file if output_file is None else output_file
    print(f'Saving to File: {output_file}')
    wb.save(output_file)
    excel_file.close()
    print('Complete')


def get_angle_data(filename,flip=False):
    df2 = pd.read_csv(filename)
    df2['t'] = df2['Frame']/df2['fps']
    df2['x_delta'] = df2['x1'] - df2['x0']
    df2['y_delta'] = df2['y1'] - df2['y0']
    df2['x_delta'] = df2['x_delta'].where((df2['x0']>df2['x1']).to_list(),df2['x1']-df2['x0'])
    df2['y_delta'] = df2['y_delta'].where((df2['x0']>df2['x1']).to_list(),df2['y1']-df2['y0'])
    df2['angle'] = np.rad2deg(np.arctan2(df2['y_delta'],df2['x_delta']))
    df2['angle'] = df2['angle'] if not flip else -df2['angle']
    df2['target_sep'] = np.sqrt(df2['x_delta']**2 + df2['y_delta']**2)
    return df2

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__,
                            formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("-e", "--excelfile", required = True, help = "dfdr excel file")
    ap.add_argument("-s", "--sheetname", default="DFDR", help = "Excel Sheet name")
    ap.add_argument("-c","--calib",required = True, help="calib file")
    ap.add_argument("-d", "--data", type=float, nargs="+", help="fold angle csv filename")
    ap.add_argument("-c","--calib_file", required=True, help="camera calibration file")
    ap.add_argument("-o","--output_file",help="Output file location")
    args = vars(ap.parse_args())